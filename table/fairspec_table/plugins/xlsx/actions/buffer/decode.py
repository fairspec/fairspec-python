from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from fairspec_table.models.data import DataRow

OFFICENS = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
TABLENS = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"


def decode_xlsx_buffer(
    data: bytes,
    *,
    format: str = "xlsx",
    sheet_name: str | None = None,
    sheet_number: int | None = None,
) -> list[DataRow]:
    if format == "ods":
        return _decode_ods(data, sheet_name=sheet_name, sheet_number=sheet_number)
    return _decode_xlsx(data, sheet_name=sheet_name, sheet_number=sheet_number)


def _decode_xlsx(
    data: bytes,
    *,
    sheet_name: str | None = None,
    sheet_number: int | None = None,
) -> list[DataRow]:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ws = _select_openpyxl_sheet(wb, sheet_name, sheet_number)
    if ws is None:
        return []

    rows: list[DataRow] = []
    for row_tuple in ws.iter_rows(values_only=True):  # ty: ignore[unresolved-attribute]
        row = list(row_tuple)
        while row and row[-1] is None:
            row.pop()
        rows.append(row)
    return rows


def _select_openpyxl_sheet(
    wb: object,
    sheet_name: str | None,
    sheet_number: int | None,
) -> object | None:
    if sheet_name is not None:
        return wb[sheet_name] if sheet_name in wb.sheetnames else None  # ty: ignore[unresolved-attribute, not-subscriptable]
    index = (sheet_number - 1) if sheet_number else 0
    if 0 <= index < len(wb.sheetnames):  # ty: ignore[unresolved-attribute]
        return wb[wb.sheetnames[index]]  # ty: ignore[unresolved-attribute, not-subscriptable]
    return None


def _decode_ods(
    data: bytes,
    *,
    sheet_name: str | None = None,
    sheet_number: int | None = None,
) -> list[DataRow]:
    from odf.opendocument import load as odf_load
    from odf.table import Table as OdfTable
    from odf.table import TableCell, TableRow

    doc = odf_load(BytesIO(data))
    sheets = doc.spreadsheet.getElementsByType(OdfTable)

    sheet = _select_ods_sheet(sheets, sheet_name, sheet_number)
    if sheet is None:
        return []

    rows: list[DataRow] = []
    for table_row in sheet.getElementsByType(TableRow):  # ty: ignore[unresolved-attribute]
        row: DataRow = []
        for cell in table_row.getElementsByType(TableCell):
            repeat_str = cell.getAttrNS(TABLENS, "number-columns-repeated")
            repeat = int(repeat_str) if repeat_str else 1
            value = _get_ods_cell_value(cell)
            row.extend([value] * min(repeat, 10000))

        while row and row[-1] is None:
            row.pop()

        rows.append(row)

    while rows and not rows[-1]:
        rows.pop()

    return rows


def _select_ods_sheet(
    sheets: list[object],
    sheet_name: str | None,
    sheet_number: int | None,
) -> object | None:
    if sheet_name is not None:
        for sheet in sheets:
            if sheet.getAttribute("name") == sheet_name:  # ty: ignore[unresolved-attribute]
                return sheet
        return None
    index = (sheet_number - 1) if sheet_number else 0
    if 0 <= index < len(sheets):
        return sheets[index]
    return None


def _get_ods_cell_value(cell: object) -> object:
    from odf.text import P

    value_type = cell.getAttrNS(OFFICENS, "value-type")  # ty: ignore[unresolved-attribute]
    if value_type == "float":
        val = cell.getAttrNS(OFFICENS, "value")  # ty: ignore[unresolved-attribute]
        float_val = float(val)
        if float_val == int(float_val):
            return int(float_val)
        return float_val
    if value_type == "boolean":
        return cell.getAttrNS(OFFICENS, "boolean-value") == "true"  # ty: ignore[unresolved-attribute]
    if value_type == "string":
        ps = cell.getElementsByType(P)  # ty: ignore[unresolved-attribute]
        if ps:
            return str(ps[0])
        return ""
    if value_type is None:
        return None
    ps = cell.getElementsByType(P)  # ty: ignore[unresolved-attribute]
    if ps:
        return str(ps[0])
    return None
