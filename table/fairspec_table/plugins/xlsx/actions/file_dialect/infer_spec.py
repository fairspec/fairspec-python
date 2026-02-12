from __future__ import annotations

from io import BytesIO

from fairspec_dataset import write_temp_file
from fairspec_metadata import CsvFileDialect, OdsFileDialect, Resource, XlsxFileDialect

from fairspec_table.plugins.xlsx.actions.buffer.encode import encode_xlsx_buffer

from .infer import infer_xlsx_file_dialect


def _create_xlsx_file(rows: list[list[object]]) -> str:
    buffer = encode_xlsx_buffer(rows, book_type="xlsx")
    return write_temp_file(buffer, format="xlsx")


class TestInferXlsxFileDialect:
    def test_should_detect_header_row_with_text_headers(self):
        path = _create_xlsx_file(
            [
                ["id", "name", "age"],
                [1, "Alice", 25],
                [2, "Bob", 30],
            ]
        )

        file_dialect = infer_xlsx_file_dialect(Resource(data=path))

        assert file_dialect == XlsxFileDialect(headerRows=[1])

    def test_should_not_detect_header_when_first_row_is_numeric(self):
        path = _create_xlsx_file(
            [
                [1, 2, 3],
                [4, 5, 6],
                [7, 8, 9],
            ]
        )

        file_dialect = infer_xlsx_file_dialect(Resource(data=path))

        assert file_dialect == XlsxFileDialect(headerRows=False)

    def test_should_detect_header_with_mixed_case_and_underscores(self):
        path = _create_xlsx_file(
            [
                ["user_id", "User_Name", "EmailAddress"],
                [1, "alice", "alice@example.com"],
                [2, "bob", "bob@example.com"],
            ]
        )

        file_dialect = infer_xlsx_file_dialect(Resource(data=path))

        assert file_dialect == XlsxFileDialect(headerRows=[1])

    def test_should_not_detect_header_with_boolean_values_in_first_row(self):
        path = _create_xlsx_file(
            [
                ["someId", 37257, 695.8, False, "2024-01-01"],
                ["anotherId", 68694, 337.73, True, "2024-01-02"],
                ["thirdId", 52019, 988.74, False, "2024-01-03"],
            ]
        )

        file_dialect = infer_xlsx_file_dialect(Resource(data=path))

        assert file_dialect == XlsxFileDialect(headerRows=False)

    def test_should_handle_single_row_files(self):
        path = _create_xlsx_file([["id", "name", "age"]])

        file_dialect = infer_xlsx_file_dialect(Resource(data=path))

        assert file_dialect == XlsxFileDialect(headerRows=False)

    def test_should_handle_empty_files(self):
        path = _create_xlsx_file([])

        file_dialect = infer_xlsx_file_dialect(Resource(data=path))

        assert file_dialect == XlsxFileDialect()

    def test_should_return_none_for_incompatible_format(self):
        resource = Resource(data="test.csv", fileDialect=CsvFileDialect())

        file_dialect = infer_xlsx_file_dialect(resource)

        assert file_dialect is None

    def test_should_return_none_for_resources_without_path(self):
        resource = Resource(data=[{"id": 1, "name": "alice"}])

        file_dialect = infer_xlsx_file_dialect(resource)

        assert file_dialect is None

    def test_should_respect_sheet_number_from_existing_dialect(self):
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws1 = wb.create_sheet("Sheet1")
        for row in [[1, 2, 3], [4, 5, 6]]:
            ws1.append(row)
        ws2 = wb.create_sheet("Sheet2")
        for row in [["id", "name"], [1, "Alice"], [2, "Bob"]]:
            ws2.append(row)
        buf = BytesIO()
        wb.save(buf)
        path = write_temp_file(buf.getvalue())

        file_dialect = infer_xlsx_file_dialect(
            Resource(data=path, fileDialect=XlsxFileDialect(sheetNumber=2))
        )

        assert file_dialect == XlsxFileDialect(headerRows=[1])

    def test_should_respect_sheet_name_from_existing_dialect(self):
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws1 = wb.create_sheet("Data")
        ws1.append([1, 2, 3])
        ws2 = wb.create_sheet("Headers")
        for row in [["id", "name"], [1, "Alice"]]:
            ws2.append(row)
        buf = BytesIO()
        wb.save(buf)
        path = write_temp_file(buf.getvalue())

        file_dialect = infer_xlsx_file_dialect(
            Resource(data=path, fileDialect=XlsxFileDialect(sheetName="Headers"))
        )

        assert file_dialect == XlsxFileDialect(headerRows=[1])

    def test_should_support_ods_format(self):
        buffer = encode_xlsx_buffer(
            [["id", "name", "age"], [1, "Alice", 25], [2, "Bob", 30]],
            book_type="ods",
        )
        path = write_temp_file(buffer)

        file_dialect = infer_xlsx_file_dialect(
            Resource(data=path, fileDialect=OdsFileDialect())
        )

        assert file_dialect == OdsFileDialect(headerRows=[1])
