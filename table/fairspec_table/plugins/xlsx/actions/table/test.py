from __future__ import annotations

from io import BytesIO


def read_test_data(path: str) -> list[dict[str, object]]:
    try:
        return _read_xlsx_test_data(path)
    except Exception:
        return _read_ods_test_data(path)


def write_test_data(
    path: str,
    rows: list[list[object]],
    *,
    sheet_number: int = 1,
    sheet_name: str | None = None,
    format: str = "xlsx",
) -> None:
    if format == "ods":
        _write_ods_test_data(path, rows, sheet_number=sheet_number, sheet_name=sheet_name)
    else:
        _write_xlsx_test_data(
            path, rows, sheet_number=sheet_number, sheet_name=sheet_name
        )


def _read_xlsx_test_data(path: str) -> list[dict[str, object]]:
    import openpyxl

    with open(path, "rb") as f:
        data = f.read()
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    headers = [str(h) for h in all_rows[0]]
    records: list[dict[str, object]] = []
    for row in all_rows[1:]:
        record: dict[str, object] = {}
        for i, h in enumerate(headers):
            record[h] = row[i] if i < len(row) else None
        records.append(record)
    return records


def _read_ods_test_data(path: str) -> list[dict[str, object]]:
    from odf.opendocument import load as odf_load
    from odf.table import Table as OdfTable
    from odf.table import TableCell, TableRow
    from odf.text import P

    OFFICENS = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    TABLENS = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"

    doc = odf_load(path)
    sheets = doc.spreadsheet.getElementsByType(OdfTable)
    if not sheets:
        return []

    sheet = sheets[0]
    all_rows: list[list[object]] = []

    for table_row in sheet.getElementsByType(TableRow):
        row: list[object] = []
        for cell in table_row.getElementsByType(TableCell):
            repeat_str = cell.getAttrNS(TABLENS, "number-columns-repeated")
            repeat = int(repeat_str) if repeat_str else 1
            value_type = cell.getAttrNS(OFFICENS, "value-type")
            value: object = None
            if value_type == "float":
                val = cell.getAttrNS(OFFICENS, "value")
                float_val = float(val)
                value = int(float_val) if float_val == int(float_val) else float_val
            elif value_type == "boolean":
                value = cell.getAttrNS(OFFICENS, "boolean-value") == "true"
            elif value_type == "string":
                ps = cell.getElementsByType(P)
                value = str(ps[0]) if ps else ""
            row.extend([value] * min(repeat, 10000))

        while row and row[-1] is None:
            row.pop()
        all_rows.append(row)

    while all_rows and not all_rows[-1]:
        all_rows.pop()

    if not all_rows:
        return []

    headers = [str(h) for h in all_rows[0]]
    records: list[dict[str, object]] = []
    for row in all_rows[1:]:
        record: dict[str, object] = {}
        for i, h in enumerate(headers):
            record[h] = row[i] if i < len(row) else None
        records.append(record)
    return records


def _write_xlsx_test_data(
    path: str,
    rows: list[list[object]],
    *,
    sheet_number: int,
    sheet_name: str | None,
) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for i in range(sheet_number):
        name = sheet_name or f"Sheet{i + 1}"
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    wb.save(path)


def _write_ods_test_data(
    path: str,
    rows: list[list[object]],
    *,
    sheet_number: int,
    sheet_name: str | None,
) -> None:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow

    from fairspec_table.plugins.xlsx.actions.buffer.encode import _create_ods_cell

    doc = OpenDocumentSpreadsheet()

    for i in range(sheet_number):
        name = sheet_name or f"Sheet{i + 1}"
        table = Table(name=name)
        for row_data in rows:
            tr = TableRow()
            for value in row_data:
                tc = _create_ods_cell(value)
                tr.addElement(tc)
            table.addElement(tr)
        doc.spreadsheet.addElement(table)

    doc.save(path)
